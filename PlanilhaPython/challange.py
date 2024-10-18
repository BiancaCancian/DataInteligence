import openpyxl
import openpyxl.workbook

#criar uma planilha(book)
book = openpyxl.Workbook()

#como vizualizar paginas 
print(book.sheetnames)

#como criar uma pagina
book.create_sheet('Computadores')

#Como selecionar uma pagina
computadores_page = book['Computadores']
computadores_page.append(['Eletronicos', 'Memoria Ram', 'Preço'])
computadores_page.append(['Comp 1', '8gb ram', 'R$7,99'])
computadores_page.append(['Comp 2', '16gb ram', 'R$4,99'])
computadores_page.append(['Comp3', '32gb ram', 'R$8,99'])

#salvar a planilha
book.save('Planilha de Computadores.xlsx')