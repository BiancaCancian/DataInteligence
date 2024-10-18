import openpyxl
import openpyxl.workbook

#criar uma planilha(book)
book = openpyxl.Workbook()

#como vizualizar paginas 
print(book.sheetnames)

#como criar uma pagina
book.create_sheet('Frutas')

#Como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '7', 'R$7,99'])
frutas_page.append(['Amora', '89', 'R$4,99'])
frutas_page.append(['Caqui', '67', 'R$8,99'])
frutas_page.append(['Marmelo', '15', 'R$5,99'])

#salvar a planilha
book.save('Planilha de Compras.xlsx')